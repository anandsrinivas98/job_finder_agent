"""
Professional Multi-Sheet Excel Generator for V2 AI Job Hunting Agent.
Generates a styled 6-sheet workbook conforming strictly to the V2 Specification.
Protects against Spreadsheet Formula Injection (CWE-1236) and validates hyperlink safety.
"""

import os
import urllib.parse
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReportGenerator:
    """Generates a professional 6-sheet Excel report with styling, hyperlinks, and filters."""

    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Typography & Palettes
        self.font_family = "Segoe UI"
        self.header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
        self.header_font = Font(name=self.font_family, size=11, bold=True, color="FFFFFF")
        self.regular_font = Font(name=self.font_family, size=10, color="1E293B")
        self.bold_font = Font(name=self.font_family, size=10, bold=True, color="0F172A")
        self.link_font = Font(name=self.font_family, size=10, color="2563EB", underline="single") # Blue Hyperlink

        self.border_thin = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        self.columns = [
            ("#", 6),
            ("Company", 24),
            ("Role", 30),
            ("Location", 22),
            ("Work Mode", 14),
            ("Posted Date & Time", 22),
            ("Match %", 12),
            ("Match Breakdown", 36),
            ("Experience", 18),
            ("Salary", 20),
            ("Job Type", 14),
            ("Source", 18),
            ("Apply Link", 16),
            ("Find Referral", 18),
            ("Find Recruiter", 18),
            ("Company Website", 18),
            ("Verification Status", 20),
            ("India Eligibility", 16),
            ("Status", 14)
        ]

    def generate_daily_report(self, jobs: List[Dict[str, Any]], custom_date: str = "") -> Path:
        """Builds and saves the complete 6-sheet Excel report."""
        date_str = custom_date or datetime.now().strftime("%Y-%m-%d")
        filename = f"Job_Report_{date_str}.xlsx"
        file_path = self.output_dir / filename

        wb = openpyxl.Workbook()
        # Remove default empty sheet
        wb.remove(wb.active)

        # 1. Sheet 1 - DAILY JOBS (All qualifying jobs)
        self._populate_sheet(wb, "DAILY JOBS", jobs)

        # 2. Sheet 2 - TOP MATCHES (Top 5-10 with Match Highlights & Apply First reason)
        top_matches = sorted(jobs, key=lambda x: float(x.get("match_score", 0)), reverse=True)[:10]
        self._populate_sheet(wb, "TOP MATCHES", top_matches, is_top_matches=True)

        # 3. Sheet 3 - REMOTE (Remote Only)
        remote_jobs = [j for j in jobs if j.get("work_mode") == "Remote" or j.get("is_remote", False)]
        self._populate_sheet(wb, "REMOTE", remote_jobs)

        # 4. Sheet 4 - AI & GENAI
        ai_jobs = [j for j in jobs if j.get("category") == "AI / ML / GenAI"]
        self._populate_sheet(wb, "AI & GENAI", ai_jobs)

        # 5. Sheet 5 - INTERNSHIPS
        intern_jobs = [j for j in jobs if j.get("job_type") == "Internship" or j.get("is_internship", False)]
        self._populate_sheet(wb, "INTERNSHIPS", intern_jobs)

        # 6. Sheet 6 - TESTING & ANALYST
        qa_analyst_jobs = [j for j in jobs if j.get("category") in ["Testing / QA", "Analyst / Entry Level"]]
        self._populate_sheet(wb, "TESTING & ANALYST", qa_analyst_jobs)

        wb.save(file_path)
        print(f"[📊 Excel Generator] Saved report with 6 sheets to: {file_path}")
        return file_path

    def _sanitize_cell_value(self, val: Any) -> Any:
        """Sanitizes text to prevent Spreadsheet Formula Injection (CWE-1236)."""
        if isinstance(val, str):
            # Prepend single quote if string begins with risky formula triggers
            if val.startswith(("=", "+", "-", "@", "\t", "\r")):
                return f"'{val}"
        return val

    def _is_safe_url(self, url: str) -> bool:
        """Validates that a URL strictly uses safe http or https schemes."""
        if not url or not isinstance(url, str):
            return False
        clean = url.strip().lower()
        return clean.startswith("http://") or clean.startswith("https://")

    def _generate_why_it_matches(self, job: Dict[str, Any]) -> str:
        """Generates a concise reason why this job is a top priority match."""
        score = float(job.get("match_score", 0))
        title = job.get("title", "")
        skills = job.get("skills", [])
        matched_skills = ", ".join(skills[:3]) if skills else "Python, APIs"
        return f"High profile match ({score}%). Direct alignment with {matched_skills} for {title}."

    def _populate_sheet(self, wb: openpyxl.Workbook, title: str, job_list: List[Dict[str, Any]], is_top_matches: bool = False):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True

        # Determine column list
        cols = list(self.columns)
        if is_top_matches:
            cols.insert(7, ("Why It Matches / Apply First", 38))

        # Header Row
        header_row = [c[0] for c in cols]
        ws.append(header_row)

        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 5, 6, 7, 11, len(cols)] else "left", vertical="center")
            cell.border = self.border_thin

        ws.row_dimensions[1].height = 28

        # Populate Data Rows
        for idx, job in enumerate(job_list, start=1):
            status = job.get("status", "NEW")
            status_badge = f"🆕 {status}" if status == "NEW" else (f"🔄 {status}" if status == "UPDATED" else f"⏳ {status}")
            company_name = job.get("company", "N/A")

            # Dynamic LinkedIn employee referral URL
            clean_comp_encoded = urllib.parse.quote(company_name.replace("'", "").replace('"', ''))
            referral_url = f"https://www.linkedin.com/search/results/people/?keywords={clean_comp_encoded}%20Software%20Engineer"

            # Dynamic Recruiter discovery URL
            rec_url = str(job.get("recruiter_linkedin") or "").strip()
            if not rec_url or rec_url in ["N/A", "none", "nan", ""]:
                rec_url = f"https://www.linkedin.com/search/results/people/?keywords={clean_comp_encoded}%20(Recruiter%20OR%20HR%20OR%20Talent%20Acquisition)"

            # Dynamic Company Website URL
            comp_url = str(job.get("company_website") or "").strip()
            if not comp_url or comp_url in ["N/A", "none", "nan", ""] or not comp_url.startswith("http"):
                comp_url = f"https://www.google.com/search?q={clean_comp_encoded}+official+website"

            # Salary fallback
            salary_val = job.get("salary")
            if not salary_val or str(salary_val).strip() in ["N/A", "nan", "none", ""]:
                salary_val = "Not Disclosed"

            # Posted date & time
            posted_val = str(job.get("posted_date") or "").strip()
            if not posted_val or posted_val in ["N/A", "none", "Date not verified"]:
                posted_val = datetime.now().strftime("%Y-%m-%d (%H:%M IST)")

            # Format match breakdown string
            mb = job.get("match_breakdown")
            if hasattr(mb, "to_summary_str"):
                mb_str = mb.to_summary_str()
            elif isinstance(mb, dict):
                mb_str = f"Skills: {mb.get('skills', 0)}/35 | Resp: {mb.get('responsibilities', 0)}/25 | Proj: {mb.get('projects', 0)}/15 | Exp: {mb.get('experience', 0)}/10"
            else:
                mb_str = "Standard Matrix"

            # Verification Status Badge
            v_status = job.get("verification_status", "UNVERIFIED")
            v_badge = f"🛡️ {v_status}" if v_status == "VERIFIED" else (f"⚠️ {v_status}" if v_status == "PARTIALLY_VERIFIED" else v_status)

            india_eligibility = "Eligible (India / Remote)" if job.get("india_eligibility_verified", True) else "Location Restricted"

            raw_row = [
                idx,
                company_name,
                job.get("title", "N/A"),
                job.get("location", "India"),
                job.get("work_mode", "N/A"),
                posted_val,
                f"{job.get('match_score', 0)}%"
            ]

            if is_top_matches:
                raw_row.append(self._generate_why_it_matches(job))

            raw_row.extend([
                mb_str,
                job.get("experience", "Fresher / 0-2 yrs"),
                salary_val,
                job.get("job_type", "Full-time"),
                job.get("source", "N/A"),
                "Apply Link ↗",
                "Ask Referral ↗",
                "Find Recruiter ↗",
                "Website ↗",
                v_badge,
                india_eligibility,
                status_badge
            ])

            row_data = [self._sanitize_cell_value(item) for item in raw_row]
            ws.append(row_data)
            row_idx = idx + 1
            ws.row_dimensions[row_idx].height = 22

            # Styling individual cells
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = self.regular_font
                cell.border = self.border_thin
                cell.alignment = Alignment(horizontal="center" if col_idx in [1, 5, 6, 7, len(cols)] else "left", vertical="center")

                # Match % styling
                if col_idx == 7:
                    score = float(job.get("match_score", 0))
                    cell.font = self.bold_font
                    if score >= 85:
                        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # light green
                    elif score >= 70:
                        cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # light yellow

                # Apply Link Hyperlink
                apply_col_idx = 14 if is_top_matches else 13
                if col_idx == apply_col_idx:
                    url = job.get("job_url", "")
                    if self._is_safe_url(url):
                        cell.value = "Apply Link ↗"
                        cell.hyperlink = url
                        cell.font = self.link_font
                    else:
                        cell.value = "N/A"

                # Find Referral LinkedIn Link
                referral_col_idx = 15 if is_top_matches else 14
                if col_idx == referral_col_idx:
                    cell.value = "Ask Referral ↗"
                    cell.hyperlink = referral_url
                    cell.font = self.link_font

                # Recruiter LinkedIn Link
                recruiter_col_idx = 16 if is_top_matches else 15
                if col_idx == recruiter_col_idx:
                    if self._is_safe_url(rec_url):
                        cell.value = "Find Recruiter ↗"
                        cell.hyperlink = rec_url
                        cell.font = self.link_font
                    else:
                        cell.value = "N/A"

                # Company Website Link
                website_col_idx = 17 if is_top_matches else 16
                if col_idx == website_col_idx:
                    if self._is_safe_url(comp_url):
                        cell.value = "Website ↗"
                        cell.hyperlink = comp_url
                        cell.font = self.link_font
                    else:
                        cell.value = "N/A"

        # Auto-adjust column widths & add auto-filter
        for col_idx, (name, width) in enumerate(cols, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(width, 12)

        if len(job_list) > 0:
            last_col_letter = get_column_letter(len(cols))
            ws.auto_filter.ref = f"A1:{last_col_letter}{len(job_list) + 1}"
            ws.freeze_panes = "A2"
