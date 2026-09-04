import pytest
import openpyxl
from pathlib import Path
from core.excel_generator import ExcelReportGenerator
from core.models import MatchBreakdown

def test_excel_generation_all_six_sheets(tmp_path):
    generator = ExcelReportGenerator(output_dir=tmp_path)
    
    mock_jobs = [
        {
            "company": "Google",
            "title": "Software Engineer Fresher",
            "location": "Bengaluru, India",
            "work_mode": "On-site",
            "posted_date": "2026-09-04",
            "match_score": 92.0,
            "match_breakdown": MatchBreakdown(skills=32, responsibilities=22, projects=14, experience=9, education=5, location=5, other=5),
            "experience": "0-1 yrs",
            "salary": "INR 18,00,000 - 24,00,000",
            "job_type": "Full-time",
            "source": "LinkedIn (JobSpy)",
            "job_url": "https://careers.google.com/jobs/results/123",
            "company_website": "https://google.com",
            "category": "Software / Development",
            "verification_status": "VERIFIED",
            "india_eligibility_verified": True,
            "status": "NEW"
        },
        {
            "company": "OpenAI",
            "title": "AI / GenAI Research Intern",
            "location": "Remote",
            "work_mode": "Remote",
            "posted_date": "2026-09-04",
            "match_score": 95.0,
            "match_breakdown": MatchBreakdown(skills=35, responsibilities=24, projects=15, experience=8, education=4, location=4, other=5),
            "experience": "Fresher",
            "salary": "$40 / hr",
            "job_type": "Internship",
            "is_internship": True,
            "is_remote": True,
            "source": "We Work Remotely",
            "job_url": "https://openai.com/careers/intern-1",
            "company_website": "https://openai.com",
            "category": "AI / ML / GenAI",
            "verification_status": "VERIFIED",
            "india_eligibility_verified": True,
            "status": "NEW"
        },
        {
            "company": "QA Tech",
            "title": "SDET / QA Automation Tester",
            "location": "Pune, India",
            "work_mode": "Hybrid",
            "posted_date": "2026-09-03",
            "match_score": 75.0,
            "match_breakdown": MatchBreakdown(skills=26, responsibilities=18, projects=12, experience=8, education=4, location=4, other=3),
            "experience": "0-2 yrs",
            "salary": "INR 6,00,000",
            "job_type": "Full-time",
            "source": "Indeed",
            "job_url": "https://qatech.com/careers/sdet",
            "company_website": "https://qatech.com",
            "category": "Testing / QA",
            "verification_status": "PARTIALLY_VERIFIED",
            "india_eligibility_verified": True,
            "status": "UPDATED"
        }
    ]

    report_path = generator.generate_daily_report(mock_jobs, custom_date="2026-09-04")
    assert report_path.exists()
    assert report_path.name == "Job_Report_2026-09-04.xlsx"

    wb = openpyxl.load_workbook(report_path)
    expected_sheets = ["DAILY JOBS", "TOP MATCHES", "REMOTE", "AI & GENAI", "INTERNSHIPS", "TESTING & ANALYST"]
    assert wb.sheetnames == expected_sheets

    # Validate DAILY JOBS sheet
    ws_daily = wb["DAILY JOBS"]
    assert ws_daily.max_row == 4 # Header + 3 records
    assert ws_daily.cell(row=2, column=2).value == "Google"
    assert ws_daily.cell(row=2, column=7).value == "92.0%"

    # Validate TOP MATCHES sheet has Why It Matches column
    ws_top = wb["TOP MATCHES"]
    header_values = [ws_top.cell(row=1, column=c).value for c in range(1, ws_top.max_column + 1)]
    assert "Why It Matches / Apply First" in header_values
