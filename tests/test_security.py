import pytest
import openpyxl
from pathlib import Path
from core.excel_generator import ExcelReportGenerator

def test_excel_formula_injection_defense(tmp_path):
    generator = ExcelReportGenerator(output_dir=tmp_path)

    malicious_jobs = [
        {
            "company": "=cmd|' /C calc'!A0",
            "title": "+2+5+cmd|' /C notepad'!A0",
            "location": "@SUM(1+1)",
            "work_mode": "\tRemote",
            "posted_date": "2026-09-03",
            "salary": "-1000",
            "experience": "0-1 yrs",
            "job_type": "Full-time",
            "job_url": "javascript:alert(1)", # Malicious URL
            "recruiter_linkedin": "file:///C:/Windows/System32/cmd.exe", # Malicious URL
            "company_website": "https://legitimate-company.com", # Valid URL
            "match_score": 88.0,
            "category": "Software / Development",
            "status": "NEW"
        }
    ]

    excel_path = generator.generate_daily_report(malicious_jobs, custom_date="test_security")
    assert excel_path.exists()

    # Load and inspect sheet cells
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["DAILY JOBS"]

    company_val = ws.cell(row=2, column=2).value
    title_val = ws.cell(row=2, column=3).value
    apply_link_val = ws.cell(row=2, column=13).value
    apply_hyperlink = ws.cell(row=2, column=13).hyperlink
    referral_val = ws.cell(row=2, column=14).value
    recruiter_val = ws.cell(row=2, column=15).value
    website_hyperlink = ws.cell(row=2, column=16).hyperlink

    # Verify formula injection triggers are escaped
    assert company_val.startswith("'=")
    assert title_val.startswith("'+")

    # Verify malicious protocols (javascript / file) are neutralized
    assert apply_hyperlink is None
    assert apply_link_val == "N/A"
    assert recruiter_val == "N/A"
    assert referral_val == "Ask Referral ↗"

    # Verify safe https link is preserved
    assert website_hyperlink is not None
    assert website_hyperlink.target == "https://legitimate-company.com"

def test_gitignore_covers_secrets():
    gitignore_path = Path(".gitignore")
    assert gitignore_path.exists()
    content = gitignore_path.read_text(encoding="utf-8")

    assert ".env" in content
    assert "*.db" in content
    assert "*.xlsx" in content
    assert "resume/*.pdf" in content
